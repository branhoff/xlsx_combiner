import openpyxl as xl
import pandas as pd
import shutil, os

from copy import copy
from os import path

def pull_naming_conventions(input_file, sheet_name):
    '''
    Takes input file with columns that list the naming conventions 
    for the different characteristics of the attribution files
    creates a dictionary where the key is the column headers and 
    the value is a list of characteristics
    '''
    df = pd.read_excel(input_file, sheet_name=sheet_name)
    columns = list(df)
    output = {}

    for i in columns:
        output[i] = df[i].tolist()
        output[i] = [x for x in output[i] if str(x) != 'nan']

    return output


def define_wksht_titles(naming_conventions):
    '''
    From naming conventions function creates dictionary of the attribution 
    file names as the key and assigns the related short variant as the value
    so we can title the wkshts in the final xlsx files to their character limits
    '''
    nc = naming_conventions

    list_of_names = [nc['Strategy Names'], nc['Attributions'], nc['Time Frames']]
    list_of_short_names = [nc['Strategy Names Shortened'], nc['Attribution Names Shortened'], nc['Time Frames Shortened']]

    output = {}
    
    for index, names in enumerate(list_of_names):
        short_names = list_of_short_names[index]
        short_titles = {names[i]: short_names[i] for i in range(len(names))}
        output = {**short_titles, **output}

    return output


def define_missing_files(categories, input_folder):
    '''

    '''
    output = []
    file_list = os.listdir(input_folder)
    file_list = [f[:-5].split('_') for f in file_list]
    
    for category in categories:
        category = category.replace('_', '')
        for index, f in enumerate(file_list):
            if category in f:
                break
            if category not in f and index + 1 == len(file_list):
                output.append(category)

    return output


def write_missing_files(filepath, missing_files):
    '''

    '''
    filepath = filepath + '/' + 'errors.txt'
    with open(filepath, 'w') as f:
        for item in missing_files:
            f.write(f'{item}\n')


def create_combined_file(filepath, categories, missing_files):
    '''

    '''
    for category in categories:
        category = category.replace('_','')
        if category in missing_files:
            continue
        else:
            wb = xl.Workbook()

            wb.save(filepath + '/' + category + '.xlsx')

def copy_worksheet(input_wb, output_wb, wksht_title):
    '''

    '''
    ws1 = input_wb.worksheets[0]

    ws2 = output_wb.create_sheet(wksht_title)

    print('running')

    for row in ws1:
        for cell in row:
            new_cell = ws2.cell(row=cell.row, column=cell.col_idx,
                    value= cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
    
    print('finished')

if __name__ == "__main__":
    nc_input = 'N:/python_scripts/FINAL/xlsx_quarterly_attribution_combine/BatcherFileNamingConventions.xlsx'
    cl_conventions = 'CL Conventions'
    mk_conventions = 'MK Conventions'

    nc = pull_naming_conventions(nc_input, mk_conventions)

    input_folder = 'N:/python_scripts/FINAL/xlsx_quarterly_attribution_combine/INPUT'

    mf = define_missing_files(nc['Strategy Names'],input_folder)

    mf_output = 'N:/python_scripts/FINAL/xlsx_quarterly_attribution_combine'

    write_missing_files(mf_output, mf)

    output_folder = 'N:/python_scripts/FINAL/xlsx_quarterly_attribution_combine/OUTPUT'

    create_combined_file(output_folder,nc['Strategy Names'],mf)

    output_files = os.listdir(output_folder)
    input_files = os.listdir(input_folder)

    wksht_titles = define_wksht_titles(nc)


    for output_item in output_files:
        output_name = '_' + output_item[:-5] + '_'
        output_wb = xl.load_workbook(output_folder + '/' + output_item)
        for input_item in input_files:
            categories = input_item[:-5].split('_')
            categories = ['_' + c + '_' for c in categories]
            if output_name in categories:
                input_wb = xl.load_workbook(input_folder + '/' + input_item)
                wksht_title = wksht_titles[categories[1]] + wksht_titles[categories[2]] + wksht_titles[categories[3]]
                copy_worksheet(input_wb, output_wb, wksht_title)
            else:
                continue
        
        output_wb.save(output_folder + '/' + output_item)
